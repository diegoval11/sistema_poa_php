import openpyxl

wb = openpyxl.Workbook()
ws = wb.active

ws['A1'] = 1      # 100%
ws['B1'] = 1      # 100%
ws['C1'] = "0%"

ws['D1'] = "=AVERAGE(A1:C1)"
wb.save('test_avg.xlsx')

