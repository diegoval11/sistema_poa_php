import openpyxl

file_path = 'storage/app/plantilla/Formulas.xlsx'
wb = openpyxl.load_workbook(file_path)
ws = wb['POA COMPLETO'] if 'POA COMPLETO' in wb.sheetnames else wb.active

month_cols = ['I', 'M', 'Q', 'V', 'Z', 'AD', 'AJ', 'AN', 'AR', 'AW', 'BA', 'BE']

for col in month_cols:
    cell = ws[f"{col}14"]
    if cell.value and isinstance(cell.value, str) and cell.value.endswith(', 0)'):
        # Replace the 0 at the end with "0%"
        new_val = cell.value[:-3] + ', "0%")'
        cell.value = new_val
        print(f"Updated {col}14: {new_val}")

wb.save(file_path)
print("Template saved successfully.")

