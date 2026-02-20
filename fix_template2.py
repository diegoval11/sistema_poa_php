import openpyxl

file_path = 'storage/app/plantilla/Formulas.xlsx'
wb = openpyxl.load_workbook(file_path)
ws = wb['POA COMPLETO'] if 'POA COMPLETO' in wb.sheetnames else wb.active

month_cols = ['I', 'M', 'Q', 'V', 'Z', 'AD', 'AJ', 'AN', 'AR', 'AW', 'BA', 'BE']

for col in month_cols:
    cell = ws[f"{col}14"]
    if cell.value and isinstance(cell.value, str):
        # Fix the double comma issue
        new_val = cell.value.replace(',, "0%")', ', "0%")')
        cell.value = new_val
        print(f"Fixed {col}14: {new_val}")

wb.save(file_path)
print("Template saved successfully.")

