import openpyxl
from openpyxl.utils import get_column_letter

wb = openpyxl.load_workbook('/home/paladin/Desktop/work/alcaldia-2.0/sistema-poa/temp/15. Gerencia de Planificación y Operaciones, llenado hasta septiembre.xlsx')
sheet = wb['POA COMPLETO'] if 'POA COMPLETO' in wb.sheetnames else wb.active

print("--- Fila 11 Formulas ---")
# Usually quarters are around N, R, V, Z or so. Let's print formulas for row 11 columns A through BG
for col_idx in range(1, 60):
    col = get_column_letter(col_idx)
    cell = sheet[f"{col}11"]
    if cell.value and str(cell.value).startswith('='):
        print(f"{col}11: {cell.value}")

print("\n--- Rows with 0 programmed ---")
# let's look for a planned activity row (say row 13 upwards) that has 0 programmed
for row in range(13, 30):
    # E is planned/unplanned usually.
    # Let's just print a few rows' programmed and real values and their %.
    # E.g. J, K, L for Jan, M, N, O for Feb... Let's just check percentages columns.
    pass
