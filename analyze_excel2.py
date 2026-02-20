import openpyxl
from openpyxl.utils import get_column_letter

wb = openpyxl.load_workbook('/home/paladin/Desktop/work/alcaldia-2.0/sistema-poa/temp/15. Gerencia de Planificación y Operaciones, llenado hasta septiembre.xlsx', data_only=False)
sheet = wb['POA COMPLETO'] if 'POA COMPLETO' in wb.sheetnames else wb.active

print("--- Headers Row 10 ---")
for col_idx in range(1, 65):
    col = get_column_letter(col_idx)
    val = sheet[f"{col}10"].value
    if val:
        val = str(val).replace('\n', ' ')
        print(f"{col}: {val}")

print("\n--- Rows 15-20 columns G-T ---")
for row in range(15, 21):
    print(f"Row {row}:")
    for col in ['E', 'G', 'H', 'I', 'J', 'K', 'L', 'M', 'N', 'O', 'P', 'Q', 'R', 'S']:
        cell = sheet[f"{col}{row}"]
        val = cell.value
        if val is not None:
            print(f"  {col}: {val}")

