import openpyxl

wb = openpyxl.load_workbook('storage/app/plantilla/Formulas.xlsx')
ws = wb['POA COMPLETO'] if 'POA COMPLETO' in wb.sheetnames else wb.active

month_cols = ['I', 'M', 'Q', 'U', 'Y', 'AC', 'AG', 'AK', 'AO', 'AS', 'AW', 'BA']
for col in month_cols:
    val = ws[f"{col}14"].value
    print(f"{col}: {val}")

