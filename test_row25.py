import openpyxl

wb = openpyxl.load_workbook('/home/paladin/Desktop/work/alcaldia-2.0/sistema-poa/temp/15. Gerencia de Planificación y Operaciones, llenado hasta septiembre.xlsx', data_only=False)
sheet = wb['POA COMPLETO'] if 'POA COMPLETO' in wb.sheetnames else wb.active

print("--- Row 25 ---")
for col in ['C', 'D', 'S', 'AF', 'AG', 'AT', 'BG', 'BH', 'BI']:
    cell = sheet[f"{col}25"]
    print(f"{col}25 (Formula): {cell.value}")

wb = openpyxl.load_workbook('/home/paladin/Desktop/work/alcaldia-2.0/sistema-poa/temp/15. Gerencia de Planificación y Operaciones, llenado hasta septiembre.xlsx', data_only=True)
sheet = wb['POA COMPLETO'] if 'POA COMPLETO' in wb.sheetnames else wb.active
print("--- Row 25 Values ---")
for col in ['C', 'D', 'S', 'AF', 'AG', 'AT', 'BG', 'BH', 'BI']:
    cell = sheet[f"{col}25"]
    print(f"{col}25 (Value): {cell.value}")

