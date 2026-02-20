import openpyxl
from openpyxl.utils import get_column_letter

wb = openpyxl.load_workbook('storage/app/plantilla/Formulas.xlsx')
ws = wb['POA COMPLETO'] if 'POA COMPLETO' in wb.sheetnames else wb.active

for c in range(1, 100):
    col = get_column_letter(c)
    val = ws[f"{col}14"].value
    if val and isinstance(val, str) and val.endswith(', 0)'):
        print(f"{col}: {val}")

