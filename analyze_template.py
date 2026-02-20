import openpyxl
from openpyxl.utils import get_column_letter

wb = openpyxl.load_workbook('/home/paladin/Desktop/work/alcaldia-2.0/sistema-poa/storage/app/plantilla/Formulas.xlsx', data_only=False)
sheet = wb['POA COMPLETO'] if 'POA COMPLETO' in wb.sheetnames else wb.active

print("--- Row 14 (Template Planificadas) ---")
for col in ['G', 'H', 'I', 'J', 'S']:
    cell = sheet[f"{col}14"]
    print(f"  {col}: {cell.value}")

